import os
import time
import datetime
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Alignment
import cv2
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

# Importar predictor
try:
    from predict_final import CaptchaPredictor
except ImportError:
    import sys
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from predict_final import CaptchaPredictor

import sys

# --- CONFIGURACIÓN DE RUTAS ---
URL_BASE = "https://apps2.mef.gob.pe/consulta-vfp-webapp/consultaCertificado.jspx"

if getattr(sys, 'frozen', False):
    base_exe = os.path.dirname(sys.executable)
    internal_res = os.path.join(base_exe, '_internal')
    BASE_RES_DIR = internal_res if os.path.exists(os.path.join(internal_res, "Models")) else base_exe
else:
    BASE_RES_DIR = os.path.dirname(os.path.abspath(__file__))

MODEL_PATH = os.path.join(BASE_RES_DIR, "Models", "custom_captcha", "model.h5")
CONFIGS_PATH = os.path.join(BASE_RES_DIR, "Models", "custom_captcha", "configs.yaml")

class MEFScraperLogic:
    def __init__(self, output_folder):
        self.driver = None
        self.output_folder = output_folder
        os.makedirs(self.output_folder, exist_ok=True)
        self.predictor = CaptchaPredictor(MODEL_PATH, CONFIGS_PATH)

    def start_browser(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def close_browser(self):
        if self.driver: self.driver.quit()

    def solve_captcha_loop(self):
        max_retries = 5
        for _ in range(max_retries):
            try:
                captcha_img = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "captchaImage")))
                timestamp = int(time.time())
                filename = os.path.join(self.output_folder, f"temp_captcha_{timestamp}.png")
                captcha_img.screenshot(filename)
                text = self.predictor.predict(filename)
                if os.path.exists(filename): os.remove(filename)
                
                input_captcha = self.driver.find_element(By.NAME, "j_captcha")
                input_captcha.clear()
                input_captcha.send_keys(text)
                return True
            except:
                try:
                    self.driver.find_element(By.ID, "captchaRef").click()
                    time.sleep(1)
                except: pass
        raise Exception("Fallo al resolver Captcha")

    def check_error_message(self):
        try:
            blockquotes = self.driver.find_elements(By.TAG_NAME, "blockquote")
            if blockquotes:
                error_text = blockquotes[0].text
                if "No se encontro" in error_text or "no existe" in error_text.lower():
                    return True, error_text
            return False, ""
        except: return False, ""

    def extract_table_data(self):
        all_dfs = []
        while True:
            try:
                WebDriverWait(self.driver, 4).until(EC.presence_of_element_located((By.ID, "certificadoDetalles")))
                html = self.driver.page_source
                dfs = pd.read_html(html, attrs={'id': 'certificadoDetalles'})
                if dfs: all_dfs.append(dfs[0])
                
                next_links = self.driver.find_elements(By.LINK_TEXT, "Next")
                if next_links:
                    next_links[0].click()
                    time.sleep(1)
                else: break
            except: break
        return pd.concat(all_dfs, ignore_index=True) if all_dfs else None

    def discover_certificates(self, sec_ejec_code, year, start_num=1, end_num=None):
        print(f"--- Escaneando UE {sec_ejec_code} (Rango: {start_num} a {end_num if end_num else 'Fin'}) ---")
        existing_numbers = []
        consecutive_empty = 0
        current_num = start_num
        max_empty_allowed = 15

        while True:
            # Condición de parada 1: Si hay un rango final y lo superamos
            if end_num and current_num > end_num:
                break
            
            # Condición de parada 2: Si no hay rango final y superamos los huecos permitidos
            if not end_num and consecutive_empty >= max_empty_allowed:
                break

            print(f"  Buscando Certificado {current_num}... (Huecos: {consecutive_empty})")
            
            # Asegurar que el driver esté listo
            try:
                self.driver.current_url
            except:
                print("Reiniciando navegador...")
                self.start_browser()

            self.driver.get(URL_BASE)
            try:
                select_year = Select(WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.ID, "anoEje"))))
                select_year.select_by_value(str(year))
                self.driver.find_element(By.ID, "secEjec").send_keys(str(sec_ejec_code))
                self.driver.find_element(By.ID, "certificado").send_keys(str(current_num))
                
                self.solve_captcha_loop()
                self.driver.find_element(By.CLASS_NAME, "button").click()
                time.sleep(0.8)

                is_error, _ = self.check_error_message()
                if is_error:
                    consecutive_empty += 1
                else:
                    existing_numbers.append(current_num)
                    consecutive_empty = 0
            except Exception as e:
                print(f"  Error en certificado {current_num}, reintentando... ({e})")
                continue # No cuenta como hueco si es error de red
            
            current_num += 1
            if current_num > 99999: break 
            
        return existing_numbers

    def process_ue(self, sec_ejec_code, year, start_num=1, end_num=None):
        print(f"\n=== PROCESANDO UE: {sec_ejec_code} ({year}) ===")
        cert_numbers = self.discover_certificates(sec_ejec_code, year, start_num, end_num)
        if not cert_numbers:
            print("No se hallaron certificados en el rango.")
            return

        all_certificates_data = []
        ue_name_official = "Desconocida"
        
        for num in cert_numbers:
            print(f" Descargando Certificado {num}...")
            self.driver.get(URL_BASE)
            try:
                select_year = Select(WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.ID, "anoEje"))))
                select_year.select_by_value(str(year))
                self.driver.find_element(By.ID, "secEjec").send_keys(str(sec_ejec_code))
                self.driver.find_element(By.ID, "certificado").send_keys(str(num))
                self.solve_captcha_loop()
                self.driver.find_element(By.CLASS_NAME, "button").click()
                time.sleep(1)
                
                if ue_name_official == "Desconocida":
                    try:
                        name_input = self.driver.find_element(By.ID, "secEjecNombre")
                        raw_name = name_input.get_attribute("value")
                        if raw_name:
                            ue_name_official = "".join([c for c in raw_name if c not in '<>:"/\\|?*']).strip()
                    except: pass

                df = self.extract_table_data()
                if df is not None:
                    df['UE_Codigo'] = sec_ejec_code
                    df['Ano'] = year
                    all_certificates_data.append(df)
            except: continue

        if all_certificates_data:
            full_df = pd.concat(all_certificates_data, ignore_index=True)
            try:
                filtered_df = full_df.iloc[:, [0, 2, 5, 6, 7, 8]].copy()
                filtered_df['Unidad_Ejecutora'] = full_df['UE_Codigo']
                filtered_df['Año'] = full_df['Ano']
                
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"Certificados_{sec_ejec_code}_{year}_{timestamp}.xlsx"
                output_path = os.path.join(self.output_folder, filename)
                
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    filtered_df.to_excel(writer, index=False, sheet_name='Certificados')
                    worksheet = writer.sheets['Certificados']
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                    alignment = Alignment(vertical='center', wrap_text=True)
                    
                    worksheet.column_dimensions['F'].width = 60 
                    max_row = len(filtered_df) + 1
                    max_col = len(filtered_df.columns)
                    
                    for r in range(1, max_row + 1):
                        if r > 1: worksheet.row_dimensions[r].height = 60
                        for c in range(1, max_col + 1):
                            cell = worksheet.cell(row=r, column=c)
                            cell.border = thin_border
                            cell.alignment = alignment
                            if r == 1: cell.fill = gray_fill

                print(f"Guardado correctamente: {output_path}")
            except Exception as e:
                print(f"Error al guardar: {e}")

def run_scraper_process(codes_list, year, output_folder, start_num=1, end_num=None):
    scraper = MEFScraperLogic(output_folder)
    scraper.start_browser()
    try:
        for code in codes_list:
            scraper.process_ue(code, year, start_num, end_num)
        print("\n=== PROCESO FINALIZADO EXITOSAMENTE ===")
    finally:
        scraper.close_browser()
